try:
    from collections import OrderedDict
except ImportError:
    from ordereddict import OrderedDict
import os
import sys
import time
import click
import polib
import openpyxl
from . import ColumnHeaders
try:
    unicode
except NameError:
    unicode = str
from pathlib import Path
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import math
from tqdm import tqdm
# openpyxl versions < 2.5.0b1
try:
    from openpyxl.cell import WriteOnlyCell
except ImportError:
    from openpyxl.writer.dump_worksheet import WriteOnlyCell

CONTEXT_SETTINGS = dict(help_option_names=['-h', '--help'])

@click.group(context_settings=CONTEXT_SETTINGS)
def poexcel():
    pass

def save(output_file, catalog):
    """Save catalog to a PO file.

    """
    with click.open_file(output_file, mode='w', encoding='utf-8') as f:
        f.write(unicode(catalog))


def po_timestamp(filename):
    local = time.localtime(os.stat(filename).st_mtime)
    offset = -(time.altzone if local.tm_isdst else time.timezone)
    return '%s%s%s' % (
        time.strftime('%Y-%m-%d %H:%M', local),
        '-' if offset < 0 else '+',
        time.strftime('%H%M', time.gmtime(abs(offset))))


@poexcel.command('fromxls',context_settings=CONTEXT_SETTINGS)
@click.option('-is', '--ignoresheet', 'ignore', multiple=True,
        type=str,
        help='Ignore sheets with specific names.')
@click.option('-od','--outdir','output_dir',
        type=click.Path(exists=True,dir_okay=True,file_okay=False), default = '.' ,
        show_default=True,
        help='output directory for po file')
@click.option('-if', '--inputfile','input_file', 
        type=click.Path(exists=True, readable=True), default = 'messages.xlsx',
        show_default=True,
        help='input xlsx file')
@click.argument('locale', required=False, nargs=-1)
def fromXLS(ignore, locale, input_file, output_dir):
    """
    Convert a XLS(X) file to a .PO file

    \b
    pet fromxls en cs
    - create en.po and cs.po files from messages.xlsx

    \b
    pet fromxls en=British.po cs=Czech.po
    - create British.po and Czech.po from message.xlsx

   \b 
    pet fromxls
    - extract all locales from messages.xlsx to appropriate po files
      in current directory
    """
    book = openpyxl.load_workbook(input_file)
    for sheet in book.worksheets:
        if ignore and str(sheet.title) in ignore:
            click.echo('Ignoring sheet: %s' % sheet.title)
            continue

        if sheet.max_row < 2:
            continue
        click.echo('Processing sheet %s' % sheet.title)
        row_iterator = sheet.iter_rows()
        headers = [c.value for c in next(row_iterator)]
        headers = dict((b, a) for (a, b) in enumerate(headers))
        msgctxt_column = headers.get(ColumnHeaders.msgctxt)
        msgid_column = headers.get(ColumnHeaders.msgid)
        tcomment_column = headers.get(ColumnHeaders.tcomment)
        comment_column = headers.get(ColumnHeaders.comment)
        occurrences_column = headers.get(ColumnHeaders.occurrences)
        if msgid_column is None:
            click.echo(u'Could not find a "%s" column' % ColumnHeaders.msgid,
                    err=True)
            continue
        if not locale:
            possible_columns = [msgctxt_column,msgid_column,tcomment_column,comment_column,occurrences_column]
            locale_first_col = max(filter(None,possible_columns)) + 1
            locale = [cell.value for cell in book['Translations'][1][locale_first_col:]]
        locales = [lc.split('=') if '=' in lc else (lc, lc+'.po') for lc in locale]
        locales = [(item[0],os.path.join(output_dir,item[1])) for item in locales]
        # print(f"locale je: {locales}")
        # print(f"outputdir je:{output_dir}")
        files_created = []
        for locale_name,file_name in tqdm(locales,desc='Po files creating:',position=0):
            # print(f"Processing locale: {locale_name}")
            msgstr_column = headers.get(locale_name)
            if msgstr_column is None:
                click.echo(u'Could not find a "%s" column' % locale_name, err=True)
                continue
            catalog = polib.POFile()
            catalog.header = u'This file was generated from %s' % input_file
            catalog.metata_is_fuzzy = True
            catalog.metadata = OrderedDict()
            catalog.metadata['PO-Revision-Date'] = po_timestamp(input_file)
            catalog.metadata['Content-Type'] = 'text/plain; charset=UTF-8'
            catalog.metadata['Content-Transfer-Encoding'] = '8bit'
            catalog.metadata['Language'] = locale_name
            catalog.metadata['Generated-By'] = 'poexceltool 0.1.0'
            for row in tqdm(sheet.iter_rows(min_row=2, max_row=sheet.max_row ), total=sheet.max_row, desc=f"creating {locale_name}", position=1, leave=False):
                # time.sleep(0.0001)
                row = [c.value for c in row]
                if not row[msgid_column]:
                    continue
                try:
                    entry = polib.POEntry(
                            msgid=row[msgid_column],
                            msgstr=row[msgstr_column] or '')
                    if msgctxt_column is not None and row[msgctxt_column]:
                        entry.msgctxt = row[msgctxt_column]
                    if tcomment_column:
                        entry.tcomment = row[tcomment_column]
                    if comment_column:
                        entry.comment = row[comment_column]
                    if occurrences_column:
                        if ':' in row[occurrences_column]:
                            entry.occurrences.append(row[occurrences_column].split(':',1))
                        else:
                            entry.occurrences.append([row[occurrences_column],''])
                    catalog.append(entry)
                except IndexError:
                    click.echo('Row %s is too short' % row, err=True)
            if not catalog:
                click.echo('No messages found, aborting', err=True)
                sys.exit(1)
            save(file_name,catalog)
            files_created.append(file_name)

    nf = len(files_created)
    click.secho(f"{nf} files created: ")
    click.echo_via_pager("\n".join( [", ".join(x) for x in [files_created[i:i+5] for i in range(0, nf, 5)] ] ))




def prepare_cell(sheet,value,font=None):
    cell = WriteOnlyCell(sheet,value)
    cell.font = font
    return cell

class CatalogFile(click.Path):
    def __init__(self):
        super(CatalogFile, self).__init__(exists=True, dir_okay=False,
                readable=True)

    def convert(self, value, param, ctx):
        if not os.path.exists(value) and '=' in value:
            # The user passed a <locale>:<path> value
            (locale, path) = value.split('=', 1)
            path = os.path.expanduser(path)
            real_path = super(CatalogFile, self).convert(path, param, ctx)
            return (locale, polib.pofile(real_path))
        else:
            real_path = super(CatalogFile, self).convert(value, param, ctx)
            catalog = polib.pofile(real_path)
            locale = catalog.metadata.get('Language')
            if not locale:
                locale = os.path.splitext(os.path.basename(real_path))[0]
            return (locale, catalog)


@poexcel.command('toxls',context_settings=CONTEXT_SETTINGS)
@click.option('-c', '--comments', multiple=True,
        type=click.Choice(['translator', 'extracted', 'reference', 'all']),
        default=['reference'],
        help='Comments to include in the spreadsheet')
@click.option('-o', '--output', type=click.File('wb'), default='messages.xlsx',
        help='Output file', show_default=True)
@click.argument('catalogs', metavar='CATALOG', nargs=-1, required=False, type=CatalogFile())
@click.option('-m', '--msgmerge', is_flag=True, help='flag for update(merge) from pot file')
def toXLS(comments, output, catalogs, msgmerge):
    """
    Convert .PO files to an XLSX file.

    \b
    guessing locale for PO files: 
        1. "Language" key in the PO metadata,
        2. filename.
    manual locale specification:
        pet toxls cs=basedir/czech/mydomain.po

    \b
    pet toxls en.po Bulgarian.po
    - add files en.po and Bulgarian.po as en and bg locale to messages.xlsx
    \b
    pet toxls en=British.po cs=Czech.po
    - add files British.po and Czech.po as en and cs locale to messages.xlsx

    \b
    pet toxls
    - add all po files from current dir to messages.xlsx
    """
    if not catalogs:
        pofiles = [f for f in Path(os.getcwd()).glob('*.po')]
        catalogs = []
        for pof in pofiles:
            catalog = polib.pofile(pof.resolve())
            locale = catalog.metadata.get('Language')
            if not locale:
                locale = pof.stem
            catalogs.append((locale,catalog))
    if msgmerge:
        potfile_path = Path(os.getcwd()).glob('*.pot').__next__()
        potfile = polib.pofile(potfile_path.resolve())
        click.secho(f'Merging with: {potfile_path.name}',italic=True )
        for (_,cat) in catalogs:
            cat.merge(potfile)
    has_msgctxt = False
    for (locale, catalog) in catalogs:
        has_msgctxt = has_msgctxt or any(m.msgctxt for m in catalog)

    if catalogs == []:
        click.secho('There is no po file to add to excel file !',fg='yellow')
        return

    messages = []
    seen = set()
    for (_, catalog) in catalogs:
        for msg in catalog:
            if not msg.msgid or msg.obsolete:
                continue
            if (msg.msgid, msg.msgctxt) not in seen:
                messages.append((msg.msgid, msg.msgctxt, msg))
                seen.add((msg.msgid, msg.msgctxt))

    book = openpyxl.Workbook(write_only=True)
    sheet = book.create_sheet(title=u'Translations')
    sheet.freeze_panes = get_column_letter(2 + (3 if 'all' in comments else len(comments) )) + '2'

    row = []
    has_msgctxt_column = has_occurrences_column = has_comment_column = has_tcomment_column = None
    if has_msgctxt:
        has_msgctxt_column = True
        row.append(ColumnHeaders.msgctxt)
    row.append(ColumnHeaders.msgid)
    if 'reference' in comments or 'all' in comments:
        has_occurrences_column = True
        row.append(ColumnHeaders.occurrences)
    if 'extracted' in comments or 'all' in comments:
        has_comment_column = True
        row.append(ColumnHeaders.comment)
    if 'translator' in comments or 'all' in comments:
        has_tcomment_column = True
        row.append(ColumnHeaders.tcomment)

    for (i, cat) in enumerate(catalogs):
        row.append(cat[0])
    
    fuzzy_font = Font(italic=True, bold=True)
    header_font = Font(sz=14 ,bold=True)
    row_widths = [len(r) for r in row]
    row_headers = [prepare_cell(sheet, r,font=header_font) for r in row]
    sheet.append(row_headers)

    ref_catalog = catalogs[0][1]

    for (msgid, msgctxt, message) in tqdm(messages,desc='Writing catalog to sheet'):
        row = []
        if has_msgctxt_column is not None:
            row.append(msgctxt)
        row.append(msgid)
        msg = ref_catalog.find(msgid, msgctxt=msgctxt)
        if has_occurrences_column:
            o = []
            if msg is not None:
                for (entry, lineno) in msg.occurrences:
                    if lineno:
                        o.append(u'%s:%s' % (entry, lineno))
                    else:
                        o.append(entry)
            row.append(u', '.join(o) if o else None)
        if has_comment_column:
            row.append(msg.comment if msg is not None else None)
        if has_tcomment_column:
            row.append(msg.tcomment if msg is not None else None)
        for cat in catalogs:
            cat = cat[1]
            msg = cat.find(msgid, msgctxt=msgctxt)
            if msg is None:
                row.append(None)
            elif 'fuzzy' in msg.flags:
                cell = prepare_cell(sheet,msg.msgstr,font=fuzzy_font)
                row.append(cell)
            else:
                row.append(msg.msgstr)
        row_widths = [ max(rw,r) for rw,r in zip( row_widths, [len(r) for r in row] ) ]
        sheet.append(row)

    book.save(output)

    # Adjusting columns width
    temp_book = openpyxl.load_workbook(output.name)
    sheet = temp_book['Translations']
    min_value = 20
    max_value = 50
    for cn in range(1,sheet.max_column+1):
        # https://www.office-forums.com/threads/xlsx-cell-width-in-twips.2163622/#post-6923644
        column_width_calculation = math.trunc((row_widths[cn-1]*7.6+5)/7*256)/256
        # print(f'column: {get_column_letter(cn)} length: {row_widths[cn-1]}, calculated length: {column_width_calculation} ')
        sheet.column_dimensions[get_column_letter(cn)].width = max( min(column_width_calculation, max_value), min_value )
    for row in tqdm(sheet.iter_rows(),total=sheet.max_row,desc='Adjusting columns width',position=0):
        for ce in tqdm(row,total=sheet.max_column,position=1,leave=False):
          ce.alignment = Alignment(wrapText=True,vertical='center')

    click.secho('Styling table ...',italic=True)
    all_data = sheet.calculate_dimension()
    # headerRowCount=0 removes filter marks from headers columns
    tab = Table(displayName="Table_TR", ref=all_data, headerRowCount=1)
    # Add a default style with striped rows
    style = TableStyleInfo(name="TableStyleMedium22", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    sheet.add_table(tab)

    temp_book.save(output.name)
    click.secho(f'{output.name} created', italic=True)


if __name__ == '__main__':
    poexcel()